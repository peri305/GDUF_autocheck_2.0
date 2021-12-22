import time
import base64
import datetime
import openpyxl
import requests
from Crypto.Cipher import PKCS1_v1_5
from Crypto.PublicKey import RSA


class YiBan:
    def __init__(self, account, passwd, address):
        self.account = account
        self.passwd = passwd
        self.address = address
        self.session = requests.session()
        self.name = ""
        self.HEADERS = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 15_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 yiban_iOS/5.0.6"
        }

    def encryptPassword(self, pwd):
        PUBLIC_KEY = '''-----BEGIN PUBLIC KEY-----
            MIICIjANBgkqhkiG9w0BAQEFAAOCAg8AMIICCgKCAgEA6aTDM8BhCS8O0wlx2KzA
            Ajffez4G4A/QSnn1ZDuvLRbKBHm0vVBtBhD03QUnnHXvqigsOOwr4onUeNljegIC
            XC9h5exLFidQVB58MBjItMA81YVlZKBY9zth1neHeRTWlFTCx+WasvbS0HuYpF8+
            KPl7LJPjtI4XAAOLBntQGnPwCX2Ff/LgwqkZbOrHHkN444iLmViCXxNUDUMUR9bP
            A9/I5kwfyZ/mM5m8+IPhSXZ0f2uw1WLov1P4aeKkaaKCf5eL3n7/2vgq7kw2qSmR
            AGBZzW45PsjOEvygXFOy2n7AXL9nHogDiMdbe4aY2VT70sl0ccc4uvVOvVBMinOp
            d2rEpX0/8YE0dRXxukrM7i+r6lWy1lSKbP+0tQxQHNa/Cjg5W3uU+W9YmNUFc1w/
            7QT4SZrnRBEo++Xf9D3YNaOCFZXhy63IpY4eTQCJFQcXdnRbTXEdC3CtWNd7SV/h
            mfJYekb3GEV+10xLOvpe/+tCTeCDpFDJP6UuzLXBBADL2oV3D56hYlOlscjBokNU
            AYYlWgfwA91NjDsWW9mwapm/eLs4FNyH0JcMFTWH9dnl8B7PCUra/Lg/IVv6HkFE
            uCL7hVXGMbw2BZuCIC2VG1ZQ6QD64X8g5zL+HDsusQDbEJV2ZtojalTIjpxMksbR
            ZRsH+P3+NNOZOEwUdjJUAx8CAwEAAQ==
            -----END PUBLIC KEY-----'''
        cipher = PKCS1_v1_5.new(RSA.importKey(PUBLIC_KEY))
        cipher_text = base64.b64encode(
            cipher.encrypt(bytes(pwd, encoding="utf8")))
        return cipher_text.decode("utf-8")

    def login(self):
        url = "https://mobile.yiban.cn/api/v4/passport/login"
        data = {
            "device": "APPLE",
            "v": "5.0.2",
            "mobile": int(self.account),
            "password": self.encryptPassword(self.passwd),
            "token": "",
            "ct": "2",
            "identify": "0",
            "sversion": "25",
            "app": "1",
            "apn": "wifi",
            "authCode": "",
            "sig": "934932a8993b5e23"
        }
        headers = {
            'Origin': 'https://mobile.yiban.cn',
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 14_7_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 yiban_iOS/5.0.2',
            'Referer': 'https://mobile.yiban.cn',
            'AppVersion': '5.0.2'
        }
        response = self.session.post(
            url=url, headers=headers, data=data).json()
        if response['response'] == 100:
            self.access_token = response['data']['access_token']
            self.HEADERS["loginToken"] = self.access_token
            return response
        else:
            raise Exception("账号或密码错误")

    def getHome(self):
        params = {
            "access_token": self.access_token,
        }
        r = self.session.get(
            url="https://mobile.yiban.cn/api/v4/home", params=params).json()
        self.name = r["data"]["user"]["userName"]
        for i in r["data"]["hotApps"]:
            if i["name"] == "易广金":
                self.url = i["url"]
        return r

    def oauth(self):
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 15_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 yiban_iOS/5.0.6",
            "Cookie": "loginToken={}; yibanM_user_token={}".format(self.access_token, self.access_token),
            "loginToken": self.access_token
        }
        data = {
            "client_id": "0b77c3ac53bd5c65",
            "redirect_uri": self.url,
        }
        oauth = self.session.post(
            url="https://oauth.yiban.cn/code/usersure", headers=headers, data=data).json()
        return oauth["reUrl"]

    def submit(self):
        auth_1 = self.session.get(
            url=self.oauth(), headers=self.HEADERS, allow_redirects=False)
        auth_2 = self.session.get(
            url=auth_1.headers["Location"], headers=self.HEADERS, allow_redirects=False)
        home = self.session.get(
            url=auth_2.headers["Location"], headers=self.HEADERS, allow_redirects=False)
        studentID = home.headers["Location"].split("studentID=")[1]
        self.session.get(
            url="https://ygj.gduf.edu.cn/Handler/device.ashx?flag=checkBindDevice", headers=self.HEADERS)
        data = {
            "flag": "save",
            "studentID": "{0}".format(studentID),
            "date": "{0}".format(get_today()),
            "health": "体温37.3℃以下（正常）",
            "address": "{0}".format(self.address),
            "isTouch": "否",
            "isPatient": "不是"
        }
        result = self.session.post(
            url="https://ygj.gduf.edu.cn/Handler/health.ashx?", headers=self.HEADERS, data=data).json()
        if int(result["code"]) == 0:
            return True
        else:
            return result["msg"]


def get_time():
    return time.strftime("[%Y-%m-%d %H:%M:%S]", time.localtime())


def get_today():
    return time.strftime("%Y-%m-%d", time.localtime())


def get_rows():
    rows = openpyxl.load_workbook('./data.xlsx').worksheets[0].max_row
    return rows


def get_data(x, y):
    cell = openpyxl.load_workbook('./data.xlsx').worksheets[0].cell(x, y).value
    return cell


def push(bark, err):
    api = "https://api.day.app/%s/易班打卡异常提醒/%s?" % (bark, err)
    send = requests.post(api)


def check(account, password, bark, address):
    t = 1
    while t <= 3:
        try:
            yb = YiBan(account, password, address)
            yb.login()
            yb.getHome()
            submit = yb.submit()
            if submit is True:
                print(get_time(), "[第%s次执行] %s 打卡成功" % (t, yb.name))
                t = 5
            else:
                print(get_time(), "{0} 打卡失败，失败原因：{1}".format(yb.name, submit))
                push(bark, "{0} 打卡失败，失败原因：{1}".format(yb.name, submit))
                t = 5
        except Exception as errors:
            if t <= 2:
                err = "[第%s次执行] %s %s" % (t, yb.account, errors)
                print(get_time(), err)
                push(bark, err)
                t += 1
            else:
                err = "[第%s次执行] %s %s" % (t, yb.account, errors)
                print(get_time(), err)
                push(bark, err)
                t = 5


if __name__ == '__main__':
    i = 2
    while i <= get_rows():
        if get_data(i, 3) is not None:
            if get_data(i, 4) is not None:
                check(get_data(i, 1), get_data(i, 2),
                      get_data(i, 3), get_data(i, 4))
            else:
                check(get_data(i, 1), get_data(i, 2),
                      get_data(i, 3), "广东省广州市天河区迎福路靠近广东金融学院")
        else:
            if get_data(i, 4) is not None:
                check(get_data(i, 1), get_data(i, 2),
                      "这里替换主（帮别人打卡的管理员的）BARK", get_data(i, 4))
            else:
                check(get_data(i, 1), get_data(i, 2),
                      "这里替换主（帮别人打卡的管理员的）BARK", "广东省广州市天河区迎福路靠近广东金融学院")
        i += 1
